"""
Run this script once to generate sample Excel template files.
Usage: python3 generate_samples.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def make_students_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    headers = ["student_id", "name", "email", "division", "password"]
    header_fill = PatternFill("solid", fgColor="185FA5")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Sample rows
    sample_rows = [
        ["2021001", "Rahul Mehta", "rahul@college.edu", "A", ""],
        ["2021002", "Priya Shah", "priya@college.edu", "B", ""],
        ["2021003", "Arjun Patel", "arjun@college.edu", "A", "mypassword123"],
    ]
    for row in sample_rows:
        ws.append(row)

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    ws.append([])
    note = ws.cell(row=ws.max_row + 1, column=1, value="NOTE: Leave password blank to use student_id as default password.")
    note.font = Font(italic=True, color="888780")

    wb.save("students_template.xlsx")
    print("✓ Created students_template.xlsx")

def make_marks_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Marks"

    headers = ["student_id", "insem1", "insem2", "insem3", "practical", "assignment", "endsem"]
    header_fill = PatternFill("solid", fgColor="185FA5")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    sample_rows = [
        ["2021001", 18, 20, "", 24, 9, 55],
        ["2021002", 16, 19, 21, 22, "", 48],
        ["2021003", 20, "", "", 25, 10, 60],
    ]
    for row in sample_rows:
        ws.append(row)

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    ws.append([])
    note = ws.cell(row=ws.max_row + 1, column=1, value="NOTE: Leave cells blank for marks not yet entered. Do NOT put 0 unless mark is actually 0.")
    note.font = Font(italic=True, color="888780")

    wb.save("marks_template.xlsx")
    print("✓ Created marks_template.xlsx")

if __name__ == "__main__":
    make_students_template()
    make_marks_template()
    print("\nDone! Use these files as templates for importing data via the Admin panel.")
